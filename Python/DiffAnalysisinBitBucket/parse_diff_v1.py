import re
import openpyxl
from openpyxl.utils import get_column_letter

def parse_diff(file_path):
    with open(file_path, 'r') as file:
        lines = file.readlines()
        
    added_lines = []
    removed_lines = []
    current_file = None

    method_pattern = re.compile(r'(public|private|protected|static|\s)+[a-zA-Z<>]+\s+\w+\(.*?\)\s*\{?')

    for line in lines:
        if line.startswith('diff --git'):
            match = re.search(r'diff --git a/(.+?) b/', line)
            if match:
                current_file = match.group(1)
        elif line.startswith('@@ '):
            continue
        elif line.startswith('+') and not line.startswith('+++'):
            method_name = find_method_name(lines, method_pattern, line)
            added_lines.append((current_file, method_name, line[1:].strip()))
        elif line.startswith('-') and not line.startswith('---'):
            method_name = find_method_name(lines, method_pattern, line)
            removed_lines.append((current_file, method_name, line[1:].strip()))

    return added_lines, removed_lines

def find_method_name(lines, method_pattern, current_line):
    for line in reversed(lines[:lines.index(current_line)]):
        if method_pattern.match(line):
            return line.strip()
    return 'Unknown Method'

def write_to_excel(added, removed):
    wb = openpyxl.Workbook()
    ws_added = wb.active
    ws_added.title = "Added Lines"
    ws_removed = wb.create_sheet(title="Removed Lines")

    headers = ["File", "Method", "Line"]

    # Writing headers
    for col_num, header in enumerate(headers, 1):
        ws_added.cell(row=1, column=col_num, value=header)
        ws_removed.cell(row=1, column=col_num, value=header)

    # Writing added lines
    for row_num, (file, method, line) in enumerate(added, 2):
        ws_added.cell(row=row_num, column=1, value=file)
        ws_added.cell(row=row_num, column=2, value=method)
        ws_added.cell(row=row_num, column=3, value=line)

    # Writing removed lines
    for row_num, (file, method, line) in enumerate(removed, 2):
        ws_removed.cell(row=row_num, column=1, value=file)
        ws_removed.cell(row=row_num, column=2, value=method)
        ws_removed.cell(row=row_num, column=3, value=line)

    # Adjust column widths
    for ws in [ws_added, ws_removed]:
        for col_num in range(1, 4):
            max_length = 0
            for cell in ws[get_column_letter(col_num)]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

    wb.save("diff_analysis.xlsx")
    print("Summary of added and removed lines with method names has been written to diff_analysis.xlsx")

added, removed = parse_diff('diff_output.txt')
write_to_excel(added, removed)
